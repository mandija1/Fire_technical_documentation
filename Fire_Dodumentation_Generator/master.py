import os
import openpyxl
from vypocet import vypocet_rizika, chunks
from what_to_print import what_to_print

''' First needs to be checked if we are in the right folder '''
os.chdir("../Fire_Dodumentation_Generator")
cwd = os.getcwd()  # This command gets the current folder path

###############################################################################
''' Reading input file '''
file = 'master_input.xlsx'  # Name of the file
wb = openpyxl.load_workbook(file, data_only=True)

in_master = wb["List1"]

Nazev_akce = in_master['B2']
Cesta = in_master['B3']
soubor = in_master['B4']
data = []  # empty list for storing data

# Vytvoří dictionary z xlsx souboru - pro hledání toho co tisknout
dim_row = in_master.max_row  # determines maximum number of rows
for row in in_master.iter_rows(min_row=7, max_row=dim_row,
                               min_col=1, max_col=3):
    for cell in row:
        data.append(cell.value)

data = list(chunks(data, 3))

for_gen = dict([(data[i][0], data[i][2]) for i in range(0, len(data))])

Nazev_akce = Nazev_akce.value
Cesta = Cesta.value
soubor = soubor.value

###############################################################################
'''Univerzální soubor k výpočtu požárního rizika '''
# soubor vytvoří i csv data pro další zpracování
[konst, wb] = vypocet_rizika(Cesta, soubor, cwd)
###############################################################################

###############################################################################
'''Co bude tištěno'''
# k definici toho co bude převedeno z výsledků do Latexu je
# stanoveno v master_input
print_input = []
for name, logi in for_gen.items():
    if logi == 1:
        print_input.append(name)
print(print_input)

for i in range(0, len(print_input)):
    what_to_print(print_input[i], Cesta, konst, soubor, wb)


###############################################################################
