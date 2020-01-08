''' Import of necessary components '''
import os
import openpyxl
import numpy as np
import pandas as pd
import csv
import sys

###############################################################################


def chunks(l, n):  # This definition splits the list into desired chunks
    # For item i in a range that is a length of l,
    for i in range(0, len(l), n):
        # Create an index range for l of n items:
        yield l[i:i+n]


def rounding(l):
    A = np.around(l, decimals=2)
    A = A.astype(float)
    return A


def vypocet_rizika(cesta, soubor):
    ''' First needs to be checked if we are in the right folder '''
    os.chdir(cesta)
    ###########################################################################
    ''' Reading input file '''
    file = soubor
    wb = openpyxl.load_workbook(file, data_only=True)

    POP = wb.get_sheet_by_name('POP')  # Opens the sheet called POP
    konst = wb.get_sheet_by_name('input konstrukce')
    ###########################################################################
    ''' Rearanging input file '''
    dim_row = POP.max_row  # determines maximum number of rows
    dim_col = POP.max_column  # determines maximum number of columns

    data = []  # empty list for storing data
    for row in POP.iter_rows(min_row=2, max_row=dim_row,
                             min_col=1, max_col=dim_col):
        for cell in row:
            data.append(cell.value)

    ###########################################################################
    ''' Výpočet součinitele a '''
    PU = wb.get_sheet_by_name('PU')  # Opens the {sheet} called POP

    ###########################################################################
    ''' Rearanging input file '''
    dim_row = PU.max_row  # determines maximum number of rows
    dim_col = PU.max_column  # determines maximum number of columns
    # print(dim_row)

    data2 = []  # empty list for storing data
    for row in PU.iter_rows(min_row=2, max_row=dim_row,
                            min_col=1, max_col=12):
        for cell in row:
            data2.append(cell.value)

    # print(data2)

    '# Assembling empty list for data storage'
    PU_data = list(chunks(data2, 12))

    # Listy pro uložení dat
    a_ni = []
    p_ni = []
    p_si = []
    c_i = []
    Nazev_PU = []
    Mistnost_PU = []
    Plocha_PU = []
    vyska_PU = []
    pol = []
    jmeno = []

    '# Arrange data into individual colomns'
    for i in range(0, len(PU_data)):
        a_ni.append((PU_data[i])[4])
        p_ni.append((PU_data[i])[5])
        p_si.append((PU_data[i])[6])
        Plocha_PU.append((PU_data[i])[2])
        Nazev_PU.append((PU_data[i])[0])
        Mistnost_PU.append((PU_data[i])[1])
        c_i.append((PU_data[i])[7])
        vyska_PU.append((PU_data[i])[3])
        pol.append((PU_data[i])[10])
        jmeno.append((PU_data[i])[11])

    # Konverze číselných listů na vektory
    a_n_array = np.asarray(a_ni)
    p_n_array = np.asarray(p_ni)
    p_s_array = np.asarray(p_si)
    a_s_array = 0.9
    c_array = np.asarray(c_i)
    Plocha_PU_array = np.asarray(Plocha_PU)
    vyska_PU_array = np.asarray(vyska_PU)

    #################################################
    '# Assembling empty list for data storage'
    POP_data = list(chunks(data, 14))

    h0 = []
    s0 = []
    Nazev = []
    SO = []
    Plocha = []
    Mistnost = []
    pocet = []
    vyska_m = []

    '# Arrange data into individual colomns'
    for i in range(0, len(POP_data)):
        h0.append((POP_data[i])[6])
        s0.append((POP_data[i])[5])
        Nazev.append((POP_data[i])[0])
        SO.append((POP_data[i])[7])
        Plocha.append((POP_data[i])[2])
        Mistnost.append((POP_data[i])[1])
        pocet.append((POP_data[i])[4])
        vyska_m.append((POP_data[i])[3])

    '# Transfers list with numbers into arrays'
    h0_array = np.asarray(h0)
    SO_array_init = np.asarray(SO)
    SO_array = np.asarray(SO)
    Plocha_array = np.asarray(Plocha)
    pocet_array = np.asarray(pocet)
    vyska_m_array = np.asarray(vyska_m)

    h0_init = np.zeros(len(h0))  # Empty array full of zeros to store data
    '# Pre-calculation'
    for i in range(0, len(h0_array)):
        # h0_init will calculate numerator for the h0 calculation
        h0_init[i] = h0_array[i] * SO_array[i] * pocet_array[i]
        # SO_array is the basis of denominator for h0 calculation
        SO_array[i] = SO_array[i] * pocet_array[i]
    for i in range(1, len(h0)):
        # Adding values over an iteration for each fire section
        if Nazev[i] == Nazev[i-1]:
            h0_init[i] += h0_init[i-1]
            SO_array[i] += SO_array[i-1]

    ###########################################################################
    ''' Defining functions for the calculations '''
    ###########################################################################


    def make_logical_hs(l, n, m):
        ''' Definiční blok pro vyhledání správných hodnot h_s
            Funkce rozlišuje mezi Názvem PÚ a popisem místnosti
            Funkce rozliší zda jsou v požárním úseku místnosti zadány vícekrát
            Podle toho přiřadí logické hodnoty tak aby se plochy místností
            neopakovaly.
            l - list Plocha_array, vyska_m_array
            n - list Nazev
            m - list Mistnost
        '''
        log = np.zeros(len(l))  # array full of zeros for data storage
        if len(l) == 1:
            if l is vyska_m_array:
                log[0] += 1
            if l is Plocha_array:
                log[0] += 1
        if len(l) > 1:
            if l is vyska_m_array:  # defining the type of calculation
                for i in range(0, len(l)):  # iterating over the range of the
                    if i == 0:  # firt iteration
                        if n[i] != n[i+1] and m[i] != m[i+1] or \
                           n[i] != n[i+1] and m[i] == m[i+1]:
                            log[i] += 1
                        else:
                            log[i] += 0
                    if i >= 1 and i < (len(l))-1:  # iteration over the inter
                        if n[i] != n[i-1] and m[i] != m[i+1]:
                            log[i] += 1
                        if n[i] == n[i-1] and m[i] != m[i+1]:
                            log[i] += 1
                    if i == (len(l)-1):  # last iteration will always be equal
                            log[i] += 1

            if l is Plocha_array:  # defining the type of calculation
                for i in range(0, len(l)):
                    if i == 0:
                        if n[i] != n[i+1]:
                            log[i] += 1
                        else:
                            log[i] += 0
                    if i >= 1 and i < (len(l))-1:
                        if n[i] != n[i+1]:
                            log[i] += 1
                    if i == (len(l)-1):
                            log[i] += 1
        return log


    def make_logical_h0(l):
        ''' Similar to tbe function make_logical_hs, it is working
            with the lists. '''
        logical = []
        if len(l) == 1:
            logical.append(1)
        if len(l) > 1:
            for i in range(0, len(l)):
                if i == 0:
                    if l[i] != l[i+1]:
                        logical.append(1)
                    else:
                        logical.append(0)
            for i in range(1, len(l)-1):
                if i >= 1 & i < len(h0)-1:
                    if l[i] == l[i-1] and l[i+1] == l[i]:
                            logical.append(0)
                    if l[i] != l[i-1] and l[i+1] == l[i]:
                            logical.append(0)
                    if l[i] != l[i-1] and l[i+1] != l[i]:
                            logical.append(1)
                    if l[i] == l[i-1] and l[i+1] != l[i]:
                            logical.append(1)
            logical.append(1)
        np.asarray(logical)
        return logical


    def uniques(seq):
        # hledá unikátní vstupy v listu
        checked = []
        for i in seq:
            if i not in checked:
                checked.append(i)
        return checked


    ##########################################################################
    ''' Names of the uniques fire section '''
    n_uniques = uniques(Nazev_PU)  # unikatni PU
    n_uniques_W = uniques(Nazev)
    # print(n_uniques_W)  # print the fire sections

    ##########################################################################
    ''' h_0 calculation '''
    log_h0 = make_logical_h0(Nazev)  # calling the function with logical output
    h0 = h0_init / SO_array  # Calculation itselfs (mean)
    h0 = h0 * log_h0  # applying logical array to find needed values
    A1 = np.nonzero(h0)  # find the indeces of nonzero values
    PU_h0 = h0[A1]  # allocate non zero values into new array
    # print('Hodnota h0:')
    # print(PU_h0)  # Print the results of h0 for each fire section

    ############################################################
    ''' Výpočet světlé výšky '''
    denom = np.zeros(len(Nazev))  # array full of zeros for data storage
    numerator_init = Plocha_array * vyska_m_array  # numerator of the h_s calc
    for i in range(0, len(Nazev)):
        # Adding Plocha values over the iteration to empty denom array
        if i == 0:
            denom[i] += Plocha[i]
        if i >= 1 and i < (len(Nazev)):
            if Nazev[i] == Nazev[i-1] and Mistnost[i] != Mistnost[i-1]:
                denom[i] += Plocha[i] + denom[i-1]
            if Nazev[i] != Nazev[i-1]:
                denom[i] += Plocha[i]
            if Nazev[i] == Nazev[i-1] and Mistnost[i] == Mistnost[i-1]:
                denom[i] = denom[i-1]

    # Calling functions for determining logical arrays
    log_num = make_logical_hs(vyska_m_array, Nazev, Mistnost)
    log_denum = make_logical_hs(Plocha_array, Nazev, Mistnost)
    # print(denom)
    denominator = log_denum * denom  # denominator
    numerator = log_num * numerator_init  # numerator of each entry

    nume_add = np.zeros(len(numerator))  # Empty array for data storage
    for i in range(0, len(numerator)):
        # adding numerators together over each fire section
        if i == 0:
            nume_add[i] += numerator[i]
        if i > 0 and i < len(numerator):
            if Nazev[i] == Nazev[i-1]:
                nume_add[i] += numerator[i] + nume_add[i-1]
            else:
                nume_add[i] += numerator[i]

    num_hs = np.zeros(len(Nazev_PU))
    den_hs = np.zeros(len(Nazev_PU))
    num_in = Plocha_PU_array * vyska_PU_array
    log_hs1 = np.zeros(len(Nazev_PU))
    for i in range(0, len(Nazev_PU)):
        if i == 0:
            den_hs[i] = Plocha_PU_array[i]
            num_hs[i] = num_in[i]
        if i > 0 and i < len(Nazev_PU):
            if Nazev_PU[i] == Nazev_PU[i-1]:
                den_hs[i] = den_hs[i-1] + Plocha_PU_array[i]
                num_hs[i] = num_hs[i-1] + num_in[i]
            if Nazev_PU[i] != Nazev_PU[i-1]:
                den_hs[i] = Plocha_PU_array[i]
                num_hs[i] = num_in[i]

    # logical h_s
    for i in range(0, len(Nazev_PU)):
        if len(Nazev_PU) > 1:
            if i == 0:
                if Nazev_PU[i] == Nazev_PU[i+1]:
                    log_hs1[i] = 0
                if Nazev_PU[i] != Nazev_PU[i+1]:
                    log_hs1[i] = 1
            if i > 0 and i < len(Nazev_PU)-1:
                if Nazev_PU[i] == Nazev_PU[i+1]:
                    log_hs1[i] = 0
                if Nazev_PU[i] != Nazev_PU[i+1]:
                    log_hs1[i] = 1
            if i == len(Nazev_PU)-1:
                log_hs1[i] = 1
        if len(Nazev_PU) == 1:
            log_hs1[i] = 1

    non_zero = np.nonzero(denominator)  # Finds the indeces of non-zero values
    non_zero_hs = np.nonzero(log_hs1)
    denominator = den_hs[non_zero_hs]  # Determine non-zero denominator values
    numerator = num_hs[non_zero_hs]  # Determine non-zero numerator values
    PU_hs = numerator / denominator  # Final calculation of h_s for each PU
    # print('Hodnota h_s:')
    # print(PU_hs)  # print the results for hs

    ###########################################################################
    ''' Interpolating from table D.1 for calculating parameter n '''
    # Plocha oken / plocha celého požárního úseku
    SO_PU = SO_array[non_zero]
    # print('Hodnota SO:')
    # print(SO_PU)
    Plocha_add = np.zeros(len(Plocha_array))

    for i in range(0, len(Plocha_add)):
        # adding numerators together over each fire section
        if i == 0:
            Plocha_add[i] += Plocha_array[i]
        if i > 0 and i < len(Plocha_array):
            if Nazev[i] == Nazev[i-1] and Mistnost[i] != Mistnost[i-1]:
                Plocha_add[i] += Plocha_array[i] + Plocha_add[i-1]
            if Nazev[i] == Nazev[i-1] and Mistnost[i] == Mistnost[i-1]:
                Plocha_add[i] += Plocha_add[i-1]
            if Nazev[i] != Nazev[i-1]:
                Plocha_add[i] += Plocha_array[i]

    Plocha_add2 = np.zeros(len(Plocha_PU_array))
    for i in range(0, len(Plocha_PU_array)):
        if i == 0:
            Plocha_add2[i] = Plocha_PU_array[i]
        if i > 0:
            if Nazev_PU[i] != Nazev_PU[i-1]:
                Plocha_add2[i] = Plocha_PU_array[i]
            if Nazev_PU[i] == Nazev_PU[i-1]:
                Plocha_add2[i] = Plocha_add2[i-1] + Plocha_PU_array[i]
    non_zero2 = []
    for i in range(0, len(Nazev_PU)):
        if len(Nazev_PU) > 1:
            if i == 0:
                if Nazev_PU[i] != Nazev_PU[i+1]:
                    non_zero2.append(i)
            if i > 0 and i < len(Nazev_PU)-1:
                if Nazev_PU[i] != Nazev_PU[i+1]:
                    non_zero2.append(i)
            if i == len(Nazev_PU)-1:
                non_zero2.append(i)
        if len(Nazev_PU) == 1:
            non_zero2.append(i)

    # print(non_zero2)
    S_PU = Plocha_add2[non_zero2]
    S_PU1 = Plocha_add

    y_label = np.array([0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.08, 0.10, 0.12,
                        0.14, 0.16, 0.18, 0.20, 0.25, 0.30, 0.35, 0.40, 0.45,
                        0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90,
                        0.95, 1.00])
    x_label = np.array([0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0])


    def find_closest(myArr, myNumber, names):
        if names == 'Dif_h' and myNumber <= 0.1:
            res1 = myArr[0]
            return res1
        if names == 'Dif_S' and myNumber <= 0.01:
            res = myArr[0]
            return res
        if names == 'S_m' and myNumber <= 5:
            res = myArr[0]
            return res
        if names == 'n' and myNumber <= 0.005:
            res = myArr[0]
            return res
        if names == 'n' and myNumber >= 0.35:
            res = 0.350
            return res
        if names == 'S_m' and myNumber == 500:
            res = 500
            return res
        if names == 'S_m' and myNumber >= 4000:
            res = 4000
            return res
        A = myArr[myArr > myNumber].min()
        B = myArr[myArr < myNumber].max()
        Log = np.isin(myArr, myNumber)
        Log = np.nonzero(Log)
        if myArr[Log] == myNumber:
            return myNumber
        else:
            return np.append(B, A)

    def interpolation_2D(matrix, idx, idy, num_h, num_S, i):
        if matrix[0, 0] == matrix[1, 0] and matrix[0, 0] == matrix[0, 1]:
            res1 = matrix[0, 0]
            return res1
        if matrix[0, 0] == matrix[0, 1] and matrix[0, 0] != matrix[1, 0]:
            A = idy[i, 1] - idy[i, 0]
            B = matrix[1, 0] - matrix[0, 0]
            C = num_S[i] - idy[i, 0]
            res2 = matrix[0, 0] + (C * B / A)
            return res2
        if matrix[0, 0] == matrix[1, 0] and matrix[0, 0] != matrix[0, 1]:
            A = idx[i, 1] - idx[i, 0]
            B = matrix[0, 1] - matrix[0, 0]
            C = num_h[i] - idx[i, 0]
            res3 = matrix[0, 0] + (C * B / A)
            return res3
        if matrix[0, 0] != matrix[0, 1] and matrix[0, 0] != matrix[1, 0]:
            A = idx[i, 1] - idx[i, 0]
            B = matrix[0, 1] - matrix[0, 0]
            C = num_h[i] - idx[i, 0]
            res01 = matrix[0, 0] + (C * B / A)
            ########################################
            A = idx[i, 1] - idx[i, 0]
            B = matrix[1, 1] - matrix[1, 0]
            C = num_h[i] - idx[i, 0]
            res02 = matrix[1, 0] + (C * B / A)
            ########################################
            A = idy[i, 1] - idy[i, 0]
            B = res02 - res01
            C = num_S[i] - idy[i, 0]
            res4 = res01 + (C * B / A)
        return res4

    len_uniques = len(n_uniques)
    len_uniques_W = len(n_uniques_W)
    idx_uniq = []
    idx_uniq_neg = []

    for i in range(0, len(n_uniques)):
        if n_uniques[i] in n_uniques_W:
            idx_uniq.append(i)
        if n_uniques[i] not in n_uniques_W:
            idx_uniq_neg.append(i)

    iii = np.asarray(idx_uniq, dtype=int)
    ooo = np.asarray(idx_uniq_neg, dtype=int)

    SO_PU_z = np.zeros(len_uniques)
    PU_h0_z = np.zeros(len_uniques)

    if len_uniques != len_uniques_W:
        if len_uniques_W == 0:
            SO_PU = SO_PU_z
            PU_h0 = PU_h0_z
        if len_uniques_W > 0:
            for i in range(0, len(iii)):
                SO_PU_z[iii[i]] = SO_PU[i]
                PU_h0_z[iii[i]] = PU_h0[i]
            for i in range(0, len(ooo)):
                SO_PU_z[ooo[i]] = SO_PU[i]
                PU_h0_z[ooo[i]] = PU_h0[i]
            SO_PU = SO_PU_z
            PU_h0 = PU_h0_z

    Dif_S = SO_PU / S_PU  # Poměr (sloupec) pro hledání hodnoty n
    Dif_h = PU_h0 / PU_hs
    n = Dif_S * (Dif_h ** (1/2))

    n1 = list(n)
    key_n = list(range(0, len_uniques))

    if len(Nazev) > 0:
        id_uniq = []
        for i in range(0, len(n_uniques)):
            if n_uniques[i] not in Nazev:
                id_uniq1 = n_uniques.index(n_uniques[i])
                id_uniq.append(id_uniq1)

        for index in sorted(id_uniq, reverse=True):
            del key_n[index]

    n_zeros = np.zeros(len_uniques)
    if len(n) < len_uniques:
        for i in range(0, len_uniques):
            if Nazev_PU[i] not in Nazev:
                n_zeros[i] = 0.005
    if len(Nazev) > 0:
        n_zeros[key_n] = n[key_n]
    n1 = n_zeros

    ky_label1 = np.array([0.005, 0.010, 0.015, 0.020, 0.025, 0.030, 0.040,
                          0.050, 0.060, 0.070, 0.080, 0.090, 0.10, 0.12, 0.14,
                          0.16, 0.18, 0.20, 0.25, 0.30, 0.35])
    kx_label1 = np.array([5, 10, 20, 30, 50, 100, 250, 500])
    kx_label2 = np.array([500, 750, 1000, 2000, 4000])

    #########################################################################
    '''Odečet tabulkové hodnoty k '''
    # Hodnoty které jsou pro odečet z tabulky nutné
    # pomocná hodnota n

    '''Převládající velikost ploch (maximální plocha místnosti PÚ)'''
    # S_m Největší plocha místnosti požárního úseku:
    S_mZero = np.zeros(len(Plocha_PU_array))
    for i in range(0, len(Plocha_PU_array)):
        if i == 0:
            S_mZero[i] = Plocha_PU_array[i]
        if i > 0:
            if Nazev_PU[i] == Nazev_PU[i-1] and \
               Plocha_PU_array[i] >= S_mZero[i-1]:
                S_mZero[i] = Plocha_PU_array[i]
            if Nazev_PU[i] == Nazev_PU[i-1] and \
               Plocha_PU_array[i] < S_mZero[i-1]:
                S_mZero[i] = S_mZero[i-1]
            if Nazev_PU[i] != Nazev_PU[i-1]:
                S_mZero[i] = Plocha_PU_array[i]
    S_m = S_mZero[non_zero2]
    # print(S_mZero)
    ####################################
    # S_m[0] = 4000
    ####################################
    # Najdi nejbližší hodnoty n a S_m
    x_idx_k = np.zeros(shape=(len(S_m), 2))
    y_idx_k = np.zeros(shape=(len(S_m), 2))
    for i in range(0, len(S_m)):
        if S_m[i] <= 500:
            x_idx_k[i] = find_closest(kx_label1, S_m[i], 'S_m')
        if S_m[i] > 500:
            x_idx_k[i] = find_closest(kx_label2, S_m[i], 'S_m')
        y_idx_k[i] = find_closest(ky_label1, n1[i], 'n')
    # print(x_idx_k)
    # print(y_idx_k)

    # Načtení tabulky
    xl = pd.ExcelFile(file)
    df = xl.parse("k pod 500")
    df1 = xl.parse("k nad 500 pod 36")
    df2 = xl.parse("k nad 500 nad 36")
    # "Vycucnutí těch správných dat (korespondujících s closest values)"
    val_k = np.zeros(shape=(len(S_m)*2, 2))
    log_k = np.zeros(len(PU_hs))

    ##################
    # Otestuj jestli je výška hs větší než 3,6
    # PU_hs[0] = 3.7
    ##################

    for i in range(0, len(S_m)):
        for m in range(0, 2):
            if m == 0:
                if S_m[i] <= 500:
                    val_k[i, m] = df.ix[y_idx_k[i, m], x_idx_k[i, m]]
                if S_m[i] > 500:
                    if PU_hs[i] < 3.6:
                        val_k[i, m] = df1.ix[y_idx_k[i, m], x_idx_k[i, m]]
                    if PU_hs[i] > 3.6:
                        val_k[i, m] = df2.ix[y_idx_k[i, m], x_idx_k[i, m]]
            if m == 1:
                if S_m[i] <= 500:
                    val_k[i, m] = df.ix[y_idx_k[i, m-1], x_idx_k[i, m]]
                if S_m[i] > 500:
                    if PU_hs[i] < 3.6:
                        val_k[i, m] = df1.ix[y_idx_k[i, m-1], x_idx_k[i, m]]
                    if PU_hs[i] > 3.6:
                        val_k[i, m] = df2.ix[y_idx_k[i, m-1], x_idx_k[i, m]]
    for i in range(0, len(S_m)):
        for n in range(0, 2):
            if n == 0:
                if S_m[i] <= 500:
                    val_k[i+len(S_m), n] = df.ix[y_idx_k[i, n+1], x_idx_k[i, n]]
                if S_m[i] > 500:
                    if PU_hs[i] < 3.6:
                        val_k[i+len(S_m), n] = df1.ix[y_idx_k[i, n+1],
                                                      x_idx_k[i, n]]
                    if PU_hs[i] > 3.6:
                        val_k[i+len(S_m), n] = df2.ix[y_idx_k[i, n+1],
                                                      x_idx_k[i, n]]
            if n == 1:
                if S_m[i] <= 500:
                    val_k[i+len(S_m), n] = df.ix[y_idx_k[i, n], x_idx_k[i, n]]
                if S_m[i] > 500:
                    if PU_hs[i] < 3.6:
                        val_k[i+len(S_m), n] = df1.ix[y_idx_k[i, n], x_idx_k[i, n]]
                    if PU_hs[i] > 3.6:
                        val_k[i+len(S_m), n] = df2.ix[y_idx_k[i, n], x_idx_k[i, n]]
    # print(val_k)
    # len(S_m)


    k_value = np.zeros(len(S_m))
    for i in range(0, len(S_m)):
        temp2 = np.zeros(shape=[2, 2])
        temp2[0] = val_k[i]
        temp2[1] = val_k[i+len(S_m)]
        k_value[i] = interpolation_2D(temp2, x_idx_k, y_idx_k, S_m, n1, i)
    # n_value[i] = interpolation_2D(temp, x_idx, y_idx, Dif_h, Dif_S, i)
    # print('K_value:')
    # print(k_value)

    ###########################################################################
    ''' Výpočet součinitele b '''
    b_in = np.zeros(len(n_uniques))
    save_i = []
    for i in range(0, len(n_uniques)):
        if n_uniques[i] not in Nazev:
            idx_hsP = n_uniques.index(n_uniques[i])
            b_in[i] = k_value[i] / (0.005 * (PU_hs[idx_hsP] ** (1/2)))
        if n_uniques[i] in Nazev:
            b_in[i] = (Plocha_add2[non_zero2[i]] * k_value[i])
            save_i.append(i)
    # (SO_array[non_zero2[i]] * h0[non_zero2[0]] ** (1/2))
    save = np.array(save_i)
    b_in[save_i] = b_in[save_i] / (SO_array[non_zero] * h0[non_zero] ** (1/2))
    # b = (S_PU1[non_zero] * k_value) / ((SO_array[A1]) * (PU_h0 ** (1/2)))
    b = b_in

    for i in range(0, len(b)):
        if b[i] > 1.7:
            b[i] = 1.7
        if b[i] < 0.5:
            b[i] = 0.5

    ###############################
    ''' Stanovení hodnoty p_n pro požární úsek s více místnostmi '''

    def mean_value(l, n, p):
        num = np.zeros(len(l))
        den = np.zeros(len(l))
        for i in range(0, len(l)):
            if i == 0:
                num[i] = p[i] * l[i]
                den[i] = l[i]
            if i > 0 and i <= len(l):
                if n[i] == n[i-1]:
                    num[i] = (p[i] * l[i]) + num[i-1]
                    den[i] = l[i] + den[i-1]
                if n[i] != n[i-1]:
                    num[i] = p[i] * l[i]
                    den[i] = l[i]
        return num, den

    log_last = np.zeros(len(Nazev_PU))
    for i in range(0, len(Nazev_PU)):
        if len(Nazev_PU) > 1:
            if i == 0:
                if Nazev_PU[i] == Nazev_PU[i+1]:
                    log_last[i] = 0
                if Nazev_PU[i] != Nazev_PU[i+1]:
                    log_last[i] = 1
            if i > 0 and i < len(Nazev_PU)-1:
                if Nazev_PU[i] == Nazev_PU[i-1] and Nazev_PU[i] == Nazev_PU[i+1]:
                    log_last[i] = 0
                if Nazev_PU[i] != Nazev_PU[i-1] and Nazev_PU[i] == Nazev_PU[i+1]:
                    log_last[i] = 0
                if Nazev_PU[i] != Nazev_PU[i-1] and Nazev_PU[i] != Nazev_PU[i+1]:
                    log_last[i] = 1
                if Nazev_PU[i] == Nazev_PU[i-1] and Nazev_PU[i] != Nazev_PU[i+1]:
                    log_last[i] = 1
            if i == len(Nazev_PU)-1:
                log_last[i] = 1
        if len(Nazev_PU) == 1:
            log_last[i] = 1

    num_p_n, den_p_n = mean_value(Plocha_PU_array, Nazev_PU, p_n_array)
    log_p_n = make_logical_h0(Nazev_PU)
    num_p_n = log_last * num_p_n
    log_num = np.nonzero(num_p_n)
    numerator_p_n = num_p_n[log_num]
    denominator_p_n = den_p_n[log_num]
    p_n = numerator_p_n / denominator_p_n  # Hodnota p_n

    ###############################
    ''' Stanovení hodnoty a_n pro požární úsek s více místnostmi '''
    num_a_n = np.zeros(len(Plocha_PU_array))
    den_a_n = np.zeros(len(Plocha_PU_array))
    for i in range(0, len(Plocha_PU_array)):
        if i == 0:
            num_a_n[i] = p_n_array[i] * Plocha_PU_array[i] * a_n_array[i]
            den_a_n[i] = Plocha_PU_array[i] * p_n_array[i]
        if i > 0 and i <= len(Plocha_PU_array):
            if Nazev_PU[i] == Nazev_PU[i-1]:
                num_a_n[i] = (p_n_array[i] * Plocha_PU_array[i] * a_n_array[i]) \
                             + num_a_n[i-1]
                den_a_n[i] = (Plocha_PU_array[i] * p_n_array[i]) + den_a_n[i-1]
            if Nazev_PU[i] != Nazev_PU[i-1]:
                num_a_n[i] = p_n_array[i] * Plocha_PU_array[i] * a_n_array[i]
                den_a_n[i] = Plocha_PU_array[i] * p_n_array[i]

    numerator_a_n = num_a_n[log_num]
    denominator_a_n = den_a_n[log_num]
    a_n = numerator_a_n / denominator_a_n  # Hodnota a_n

    ###############################
    ''' Hodnota stálého požárního zatížení'''
    num_p_s, den_p_s = mean_value(Plocha_PU_array, Nazev_PU, p_s_array)
    numerator_p_s = num_p_s[log_num]
    denominator_p_s = den_p_s[log_num]
    p_s = numerator_p_s / denominator_p_s  # Hodnota p_s

    ###############################
    ''' Součinitel c'''
    c = c_array[log_num]

    ###############################
    ''' Součinitel a pro celý požární úsek '''
    a = ((p_n * a_n) + (p_s * a_s_array)) / (p_n + p_s)

    ###############################
    ''' Požárně výpočtové zatížení '''
    p = (p_n + p_s) * a * b * c
    p_only = (p_n + p_s)

    ###############################
    ''' Ověření soustředěného požárního zatížení '''
    # Spočtení požárního zatížení p pro každý prostor
    p_array = p_n_array + p_s_array
    # print(p_array)

    # Součinitel a pro každou místnost
    a_m = ((p_n_array * a_n_array) + (p_s_array * a_s_array)) /\
          (p_n_array + p_s_array)
    # print(a_m)

    # Spočtení p.a pro celý požární úsek
    p_PU = p_n + p_s
    # L_cond = 2 * (p_PU * a)
    R_cond = 50

    # Součin p.a pro danou místnost
    pa_m = a_m * p_array

    def uniques_idx(Name, uniques):
        log = []
        for item in uniques:
            if Name.count(item) > 1:
                indices = [i for i, x in enumerate(Name) if x == item]
                log.append(indices)
        return log


    log_cond = uniques_idx(Nazev_PU, n_uniques)
    sum_L = sum(isinstance(i, list) for i in log_cond)
    Left_p = []
    Left_a = []
    if sum_L > 0:
        for i in range(0, len(log_cond)):
            # print(len(log_cond[i]))
            if len(log_cond[i]) > 2:
                for k in range(0, len(log_cond[i])):
                    temp = [x for n, x in enumerate(log_cond[i]) if n != k]
                    num_p_temp, den_p_temp = mean_value(Plocha_PU_array[temp],
                                                        Nazev_PU, p_array[temp])
                    num_a_temp, den_a_temp = mean_value(Plocha_PU_array[temp],
                                                        Nazev_PU, a_m[temp])
                    new_num_p = np.sum(num_p_temp)
                    new_den_p = np.sum(den_p_temp)
                    new_num_a = np.sum(num_a_temp)
                    new_den_a = np.sum(den_a_temp)
                    p_con = new_num_p / new_den_p
                    a_con = new_num_a / new_den_a
                    Left_p.append(p_con)
                    Left_a.append(a_con)

            if len(log_cond[i]) == 2:
                for k in range(0, len(log_cond[i])):
                    temp = [x for n, x in enumerate(log_cond[i]) if n != k]
                    p_con = p_array[temp]
                    a_con = a_m[temp]
                    Left_p.append(p_con)
                    Left_a.append(a_con)
    Left_p = np.array(Left_p)
    Left_a = np.array(Left_a)
    Left_cond = Left_p * Left_a * 2

    log_cond2 = []
    for i in range(0, sum_L):
        for k in range(0, len(log_cond[i])):
            log_cond2.append(log_cond[i][k])

    Mid_con = pa_m[log_cond2]
    # print(log_cond2)

    log_pa = [0] * len(log_cond2)
    for i in range(0, len(log_cond2)):
        if Left_cond[i] < Mid_con[i] > R_cond and \
           Plocha_PU_array[log_cond2][i] >= 25:
            log_pa[i] = 1
    # print(log_pa)
    # print(a)

    # Najdi místnosti, které mají soustředěné požární zatížení
    log_idx = [0] * len(log_pa)
    for i in range(0, len(log_pa)):
        if log_pa[i] == 1:
            log_idx[i] = log_cond2[i]

    log_idx = list(filter((0).__ne__, log_idx))
    # Hodnota a pro požární úsek s SPZ
    a_m_s = a_m[log_idx]
    # Hodnota p pro požární úsek s SPZ
    p_s_s = p_s_array[log_idx]
    p_n_s = p_n_array[log_idx]
    # Hodnota c pro požární úsek s SPZ
    c_array = c_array[log_idx]


    # Hodnota součinitele b pro požární úsek s mísntosti b
    log_idx_spz = []
    for i in range(0, len(n_uniques)):
        if len(log_idx) >= 1:
            for k in range(0, len(log_idx)):
                if Nazev_PU[log_idx[k]] == n_uniques[i]:
                    log_idx_spz.append(i)

    # print(log_idx_spz)  # index pro změnu ve výsledcích

    # Požárně výpočtové zatížení pro místnost s SPZ
    p_spz = (p_s_s + p_n_s) * a_m_s * b[log_idx_spz] * c_array
    for i in range(0, len(log_idx_spz)):
        p[log_idx_spz[i]] = p_spz[i]  # Nahrazení starého zatížení novým
    # print('Pozarni zatizeni p:')
    # print(p)
    if len(p_spz) > 0:
        for i in range(0, len(log_idx)):
            print()
            print('!!! Pozor !!! soustredene pozarni zatizeni v PU ' +
                  str(n_uniques[log_idx_spz[i]]) + ' v mistnosti ' +
                  str(Mistnost_PU[log_idx[i]]))
            print()

    ###########################################################
    ###########################################################
    ''' Zatřídění do stupně požární bezpenčosti'''
    # Volání proměnných
    Info = wb.get_sheet_by_name('Info')  # Opens the sheet called Info

    type_sys = (Info['D4']).value  # Načtení konstrukčního systému
    h_p = (Info['D3']).value  # Načtení požární výšky objektu
    podlazi = (Info['D5']).value
    vytah = (Info['H3']).value
    vytah_typ = (Info['I3']).value
    sachta = (Info['H4']).value
    sachta_typ = (Info['I4']).value

    os.chdir("c:/Users/Honza/Google Drive/Work/Generator_zprav/minor/")
    sys.path.insert(0, "c:/Users/Honza/Google Drive/Work/Generator_zprav/minor/")
    from stupen import spb_def, rozmer
    SPB_result = spb_def(h_p, type_sys, p, a, podlazi, n_uniques)

    ###############################################################################
    ''' Ověření maximálních rozměrů '''
    [delka_PU, sirka_PU] = rozmer(a, type_sys, podlazi, h_p)

    # print(delka_PU, sirka_PU)

    # print(p_n)
    # print(a_n)

    # for i in range(0, len(a)):
        # print(n_uniques[i], p_only[i], S_PU[i], p_only[i] * S_PU[i])
        # print(n_uniques[i], a_round[i], delka_PU[i], sirka_PU[i])
        # print(n_uniques[i], S_PU[i], a_round[i], 0.15 * ((S_PU[i] * a[i]) ** (1/2)))
        # print(n_uniques[i], p_n[i], a_n[i])

    ##########################################################################
    ''' Výsledky a export'''
    dire = 'data'
    cesta2 = cesta
    for i in range(0, len(dire)):
        cesta2 += dire[i]
    os.chdir(cesta2)

    # Round a values
    p_round = rounding(p)
    a_round = rounding(a)
    b_round = rounding(b)
    c_round = rounding(c)

    log_pol = []
    for i in range(0, len(log_p_n)):
        if log_p_n[i] == 1:
            log_pol.append(i)

    pol1 = []
    for item in log_pol:
        pol1.append(pol[item])

    results = []
    for i in range(0, len(p)):
        results.append(n_uniques[i])
        results.append(S_PU[i])
        results.append(a_round[i])
        results.append(b_round[i])
        results.append(c_round[i])
        results.append(p_round[i])
        results.append(SPB_result[i])
        results.append(p_only[i])
        results.append(SO_PU[i])
        results.append(PU_hs[i])
        results.append(PU_h0[i])
        results.append(p_only[i])
        results.append(p_n[i])
        results.append(a_n[i])

    for i in range(0, len(n1)):
        if n1[i] == 0:
            n1[i] = 0.00500

    for i in range(0, len(results)):
        if isinstance(results[i], np.float64):
            results[i] = '%.2f' % results[i]
        else:
            results[i] = results[i]
    list_res = list(chunks(results, 14))

    for i in range(0, len(p)):
        list_res[i].append(n1[i])
        list_res[i][-1] = '%.5f' % list_res[i][-1]
        list_res[i].append(k_value[i])
        list_res[i][-1] = '%.5f' % list_res[i][-1]

    J_id = ([i for i, e in enumerate(log_p_n) if e != 0])
    Jmeno = []
    for item in J_id:
        Jmeno.append(jmeno[item])

    for i in range(0, len(p)):
        list_res[i].append(Jmeno[i])
        list_res[i].append(float(delka_PU[i]))
        list_res[i].append(float(sirka_PU[i]))
        list_res[i].append(pol1[i])
        list_res[i].append(p_s[i])
    
    '''with open('results.csv', 'w') as file:
        output = csv.writer(file, delimiter='\t')
        output.writerows(list_res)'''

    with open('results.csv', 'w', newline='', encoding='utf-8') as file:
        output = csv.writer(file, dialect='excel', delimiter='\t')
        output.writerows(list_res)

    ###########################################################################
    '''Generování dat pro appendix - zadání PU'''
    # Shromáždění dat pro
    raw_data_PU = []
    for i in range(0, len(Nazev_PU)):
        raw_data_PU.append(Nazev_PU[i])
        raw_data_PU.append(Mistnost_PU[i])
        raw_data_PU.append(format(Plocha_PU[i], '.2f'))
        raw_data_PU.append(format(vyska_PU[i], '.2f'))
        raw_data_PU.append(format(a_ni[i], '.2f'))
        raw_data_PU.append(format(p_ni[i], '.2f'))
        raw_data_PU.append(format(p_si[i], '.2f'))
        raw_data_PU.append(format(c_i[i], '.2f'))
        if pol[i] == 'B.1 pol.10 OB2':
            pol[i] = 'B.1 pol.10'
        raw_data_PU.append(pol[i])
    raw_data_PU2 = list(chunks(raw_data_PU, 9))

    # Zapsání do souboru csv
    with open('raw_data_PU.csv', 'w', newline='', encoding='utf-8') as file:
        output = csv.writer(file, dialect='excel', delimiter='\t')
        output.writerows(raw_data_PU2)

    ###############################################################################
    '''Generování dat pro appendix - zadání POP'''
    # Shromáždění dat pro
    raw_data_POP = []
    for i in range(0, len(Nazev)):
        raw_data_POP.append(Nazev[i])
        raw_data_POP.append(Mistnost[i])
        raw_data_POP.append(pocet[i])
        raw_data_POP.append(format(s0[i], '.2f'))
        raw_data_POP.append(format(h0_array[i], '.2f'))
        raw_data_POP.append(format(SO[i], '.2f'))
    raw_data_POP = list(chunks(raw_data_POP, 6))

    with open('raw_data_POP.csv', 'w', newline='', encoding='utf-8') as file:
        output = csv.writer(file, dialect='excel', delimiter='\t')
        output.writerows(raw_data_POP)

    #######################################################################
    '''Generování dat pro appendix - zadání info'''
    # Shromáždění dat pro
    raw_data_info = []
    h_p = [h_p]
    podlazi = [podlazi]
    vytah = [vytah]
    sachta = [sachta]
    type_sys = [type_sys]
    vytah_typ = [vytah_typ]
    sachta_typ = [sachta_typ]
    raw_data_info.append(type_sys)
    raw_data_info.append(h_p)
    raw_data_info.append(list(podlazi))
    raw_data_info.append(vytah)
    raw_data_info.append(sachta)
    raw_data_info.append(vytah_typ)
    raw_data_info.append(sachta_typ)

    with open('raw_data_info.csv', 'w', newline='', encoding='utf-8') as file:
        output = csv.writer(file, dialect='excel', delimiter='\t')
        output.writerows(raw_data_info)
    return konst
