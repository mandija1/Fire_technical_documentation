import sys
import os
import openpyxl
from pylatex import Document, NoEscape, Section, Itemize
from pylatex.base_classes.command import Options
sys.path.insert(0, '../Fire_Dodumentation_Generator/minor/')
from konst_data_prep import konst_data_prep
from read_results import read_results
import numpy as np


def cihlyPO(b, interpolate, i, myArr):
    '''Určení lokace v tabulce'''
    if b[i] < interpolate[0]:
        raise NameError('!!! Mimo rozsah tabulkek !!!')
    if b[i] == interpolate[0]:
        min1 = interpolate[0]
    elif b[i] in interpolate:
        min1 = b[i]
    elif b[i] < interpolate[5] and b[i] not in interpolate:
        min1 = myArr[myArr < b[i]].max()
    elif b[i] >= interpolate[5]:
        min1 = interpolate[5]

    '''if b[i] >= interpolate[5]:
        max1 = interpolate[5]
    elif b[i] <= interpolate[0]:
        max1 = interpolate[0]
    elif b[i] > interpolate[0]:
        max1 = myArr[myArr > b[i]].min()'''

    '''Určení požární odolnosti'''
    min_id = interpolate.index(min1)
    idx = []
    for k in range(0, len(interpolate)):
        if min1 == interpolate[k]:
            idx.append(k)
    idd = max(idx)
    if idd == 0:
        check = 30
    if idd == 1:
        check = 45
    if idd == 2:
        check = 60
    if idd == 3:
        check = 90
    if idd == 4:
        check = 120
    if idd == 5:
        check = 180
    return check


geometry_options = {
    "margin": "0.5cm",
    "includeheadfoot": True
    }
doc = Document(page_numbers=True, geometry_options=geometry_options)
doc.preamble.append(NoEscape(r'\usepackage[czech]{babel}'))
doc.documentclass.options = Options('10pt')


def E1_generator(soubor, cesta, data_konst, vystup_dir, data_dir):
    os.chdir(data_dir)
    data_PU = read_results('results.csv')
    info1 = read_results('raw_data_info.csv')
    podlazi = int(info1[2][0])
    info = info1[0][0]
    SPB_check = []
    n_uniques = []
    d = {'I.': 1, 'II.': 2, 'III.': 3, 'IV.': 4, 'V.': 5, 'VI.': 6, 'VII.': 7}
    e = {1: 'I.', 2: 'II.', 3: 'III.', 4: 'IV.', 5: 'V.', 6: 'VI.', 7: 'VII.'}
    check_PP = {'I.': 30, 'II.': 45, 'III.': 60, 'IV.': 90, 'V.': 120,
                'VI.': 180, 'VII.': 180}
    check_NP = {'I.': 15, 'II.': 30, 'III.': 45, 'IV.': 60, 'V.': 90,
                'VI.': 120, 'VII.': 180}
    for i in range(0, len(data_PU)):
        SPB_check.append(data_PU[i][6])
        n_uniques.append(data_PU[i][0])

    os.chdir(cesta)
    [Nazev_kce, Specif1, Specif2, b, a,
     h, Ly, Lx, zeb_a] = konst_data_prep('E.1 PKD_Stěna', data_konst)
    wb = openpyxl.load_workbook(soubor, data_only=True)
    Data_PKD = wb.get_sheet_by_name('Databaze_PKD')

    Name_data = []
    prep_data = []
    dim_row = Data_PKD.max_row  # determines maximum number of rows
    for row in Data_PKD.iter_rows(min_row=4, max_row=dim_row,
                                  min_col=1, max_col=1):
        for cell in row:
            Name_data.append(cell.value)
    for row in Data_PKD.iter_rows(min_row=4, max_row=dim_row,
                                  min_col=3, max_col=3):
        for cell in row:
            prep_data.append(cell.value)

    preidx = Name_data.index('Porotherm')
    Name_data_prep = Name_data[preidx:]
    Prep_data_text = prep_data[preidx:]
    Name_data = Name_data[:preidx]
    prep_data = prep_data[:preidx]

    dictionary = dict(zip(Name_data, prep_data))
    dictionary_text = dict(zip(Name_data_prep, Prep_data_text))
    with doc.create(Section('Zhodnocení požární odolnosti konstrukcí')):
        spb_val = []
        for i in range(0, len(SPB_check)):
            spb_val.append(d[SPB_check[i]])
        spb_max = max(spb_val)
        spb_max = e[spb_max]
        if spb_max == 'I.':
            if info == 'nehořlavý':
                with doc.create(Itemize()) as itemize:
                    itemize.add_item("Normový požadavek na požární stěny v NP - I.SPB - (R)EI 15 DP1")
                for i in range(0, len(n_uniques)):
                    if 'P' in n_uniques[i]:
                        itemize.add_item("Normový požadavek na požární stěny v PP - I.SPB - (R)EI 30 DP1")
                    break
                if podlazi > 1:
                    itemize.add_item("Normový požadavek na požární stěny v posledním NP - III.SPB - (R)EI 30 DP1")
        if spb_max == 'II.':
            with doc.create(Itemize()) as itemize:
                itemize.add_item("Normový požadavek na požární stěny v NP - II.SPB - (R)EI 30 DP1")
        if spb_max == 'III.':
            if info == 'nehořlavý':
                with doc.create(Itemize()) as itemize:
                    itemize.add_item("Normový požadavek na požární stěny v NP - III.SPB - (R)EI 45 DP1")
                    for i in range(0, len(n_uniques)):
                        if 'P' == n_uniques[i][0]:
                            itemize.add_item("Normový požadavek na požární stěny v PP - III.SPB - (R)EI 60 DP1")
                        break
                if podlazi > 1:
                    itemize.add_item("Normový požadavek na požární stěny v posledním NP - III.SPB - (R)EI 30 DP1")
        if spb_max == 'IV.':
            with doc.create(Itemize()) as itemize:
                itemize.add_item("Normový požadavek na požární stěny v NP - IV.SPB - (R)EI 60 DP1")
        if spb_max == 'V.':
            with doc.create(Itemize()) as itemize:
                itemize.add_item("Normový požadavek na požární stěny v NP - V.SPB - (R)EI 90 DP1")
        if spb_max == 'VI.':
            with doc.create(Itemize()) as itemize:
                itemize.add_item("Normový požadavek na požární stěny v NP - VI.SPB - (R)EI 120 DP1")
        if spb_max == 'VII.':
            with doc.create(Itemize()) as itemize:
                itemize.add_item("Normový požadavek na požární stěny v NP - VII.SPB - (R)EI 180 DP1")

        save_var = []
        save_b = []
        for i in range(0, len(Specif2)):
            if Specif2[i] in dictionary:
                if Specif2[i] == 'ŽB. nenosné stěny':
                    '''TempName = Name_data.index(Specif2[i])
                    if TempName > 42:
                        doc.append(dictionary[Specif2[i]])'''
                    ''' Posouzení podle tab. 2.2 - nenosné stěny '''
                    # if 0 <= TempName < 2:
                    if b[i] < 60:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    if 60 <= b[i] < 70:
                        check = 30
                    if 70 <= b[i] < 80:
                        check = 45
                    if 80 <= b[i] < 100:
                        check = 60
                    if 100 <= b[i] < 120:
                        check = 90
                    if 120 <= b[i] < 150:
                        check = 120
                    if b[i] >= 150:
                        check = 180
                    token = 'železobetonovými nenosnými stěnami'
                    token_tab = '2.2'
                    token_type = 'EI'
                ''' Posouzení podle tabulky 2.3 - nosné stěny '''
                if Specif2[i] == 'ŽB. nosné stěny':
                    if b[i] < 120 and a[i] < 10:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    if 120 <= b[i] < 125 and a[i] < 25:
                        check = 30
                    if 125 <= b[i] < 130 and a[i] < 25:
                        check = 45
                    if (130 <= b[i] < 140 and a[i] < 25)\
                       or (b[i] > 140 and a[i] < 25):
                        check = 60
                    if 140 <= b[i] < 160 and 25 <= a[i] < 35\
                       or (b[i] > 160 and a[i] < 35):
                        check = 90
                    if 160 <= b[i] < 210 and 35 <= a[i] < 50\
                       or (b[i] > 210 and a[i] < 50):
                        check = 120
                    if b[i] >= 210 and a[i] >= 50:
                        check = 180
                    token = 'železobetonovými nosnými stěnami'
                    token_tab = '2.3'
                    token_type = 'REI'
                if Specif2[i] == 'Nenosné (všechny skupiny) – 500<p<2400':
                    top = [100, 100, 100, 140, 170, 190]
                    bottom = [70, 70, 70, 100, 140, 140]
                    interpolate = []
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-500)* ((bottom[n]-top[n])/(2400-500))
                        interpolate.append(intt)
                        if interpolate[n] - b[i] >= 0:
                            check_len = len(interpolate)
                            if check_len == 1:
                                check = 60
                            if check_len == 4:
                                check = 90
                            if check_len == 5:
                                check = 120
                            if check_len == 6:
                                check = 180
                            break
                        if len(interpolate) == 6:
                            interpolate.append(1)
                    token = 'z nenosných pálených cihel'
                    token_tab = '6.1.1'
                    token_type = 'EI'

                if Specif2[i] == 'Skupina 1S – 1000<p<2400':
                    interpolate = []
                    top = [90, 90, 90, 100, 140, 190]
                    bottom = [90, 90, 90, 90, 140, 140]
                    top_rho = 2400
                    bottom_rho = 1000
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pálených zdících prvků skupiny 1S'
                    token_tab = '6.1.2, číslo řádku 1.1 respektive 1.2'
                    token_type = 'REI'

                if Specif2[i] == 'Skupina 1 – 800<p<2400':
                    interpolate = []
                    top = [100, 100, 100, 170, 170, 190]
                    bottom = [90, 90, 90, 90, 140, 170]
                    top_rho = 2400
                    bottom_rho = 800
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pálených zdících prvků skupiny 1'
                    token_tab = '6.1.2, číslo řádku 2.1 respektive 2.2'
                    token_type = 'REI'

                if Specif2[i] == 'Skupina 1 – 500<p<800':
                    top = [100, 200, 200, 200, 365, 365]
                    bottom = [100, 170, 170, 170, 300, 300]
                    top_rho = 800
                    bottom_rho = 500
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pálených zdících prvků skupiny 1'
                    token_tab = '6.1.2, číslo řádku 2.3 respektive 2.4'
                    token_type = 'REI'

                if Specif2[i] == 'Skupina 2 – 800<p<2200, ct>25%':
                    interpolate = []
                    top = [100, 100, 100, 170, 240, 240]
                    bottom = [100, 100, 100, 140, 140, 240]
                    top_rho = 2200
                    bottom_rho = 800
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pálených zdících prvků skupiny 2'
                    token_tab = '6.1.2, číslo řádku 3.1 respektive 3.2'
                    token_type = 'REI'

                if Specif2[i] == 'Skupina 2 – 700<p<800, ct>25%':
                    interpolate = [100, 100, 170, 240, 300, 365]
                    top_rho = 800
                    bottom_rho = 700
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pálených zdících prvků skupiny 2'
                    token_tab = '6.1.2, číslo řádku 3.4'
                    token_type = 'REI'

                if Specif2[i] == 'Skupina 2 – 500<p<900, 16%<ct<25%':
                    interpolate = [100, 170, 170, 240, 300, 365]
                    top_rho = 900
                    bottom_rho = 500
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pálených zdících prvků skupiny 2'
                    token_tab = '6.1.2, číslo řádku 3.6'
                    token_type = 'REI'

                if Specif2[i] == 'Skupina 3 – 500<p<800':
                    interpolate = [100, 200, 240, 300, 365, 425]
                    top_rho = 1200
                    bottom_rho = 500
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pálených zdících prvků skupiny 3'
                    token_tab = '6.1.2, číslo řádku 4.2'
                    token_type = 'REI'

                if Specif2[i] == 'Skupina 4 – 500<p<800':
                    interpolate = [240, 240, 240, 300, 365, 425]
                    top_rho = 1200
                    bottom_rho = 500
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pálených zdících prvků skupiny 4'
                    token_tab = '6.1.2, číslo řádku 5.2'
                    token_type = 'REI'

                if Specif2[i] == 'Neznámé cihly':
                    interpolate = [240, 240, 240, 300, 365, 425]
                    top_rho = 1200
                    bottom_rho = 500
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pálených zdících prvků skupiny 4'
                    token_tab = '6.1.2, číslo řádku 5.2'
                    token_type = 'REI'

                if Specif2[i] == 'Vápenopískové (nenosné) 600<p<2400':
                    interpolate = []
                    top = [70, 90, 90, 100, 140, 170]
                    bottom = [50, 70, 70, 100, 140, 140]
                    top_rho = 2400
                    bottom_rho = 600
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z nenosných vápenopískových cihel'
                    token_tab = '6.2.1, číslo řádku 1.1 respektive 1.2'
                    token_type = 'EI'

                if Specif2[i] == 'Vápenopískové – Skupina 1S – 1700<p<2400':
                    interpolate = []
                    top = [90, 90, 90, 100, 170, 170]
                    bottom = [90, 90, 90, 100, 140, 170]
                    top_rho = 2400
                    bottom_rho = 1700
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z vápenopískových zdících prvků skupiny 1S'
                    token_tab = '6.2.2, číslo řádku 1.1 respektive 1.2'
                    token_type = 'REI'

                if Specif2[i] == 'Vápenopískové – Skupina 1 – 1400<p<2400':
                    interpolate = []
                    top = [100, 100, 100, 100, 200, 240]
                    bottom = [100, 100, 100, 100, 140, 190]
                    top_rho = 2400
                    bottom_rho = 1400
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z vápenopískových zdících prvků skupiny 1'
                    token_tab = '6.2.2, číslo řádku 2.1 respektive 2.2'
                    token_type = 'REI'

                if Specif2[i] == 'Vápenopískové – Skupina 2 – 700<p<1600':
                    interpolate = []
                    top = [100, 100, 100, 100, 200, 240]
                    bottom = [100, 100, 100, 100, 140, 190]
                    top_rho = 1600
                    bottom_rho = 700
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z vápenopískových zdících prvků skupiny 2'
                    token_tab = '6.2.2, číslo řádku 3.1 respektive 3.2'
                    token_type = 'REI'

                if Specif2[i] == 'Beton nenosné skupina 1 pórovité kamenivo 400<p<1600':
                    interpolate = []
                    top = [50, 70, 90, 140, 140, 140]
                    bottom = [50, 50, 70, 70, 140, 140]
                    top_rho = 1600
                    bottom_rho = 400
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových tvárnic s pórovitým kamenivem skupiny 1'
                    token_tab = '6.3.1, číslo řádku 1.1 respektive 1.2'
                    token_type = 'EI'

                if Specif2[i] == 'Beton nenosné skupina 1 hutné kamenivo 1400<p<2000':
                    interpolate = []
                    top = [50, 70, 90, 140, 140, 190]
                    bottom = [50, 50, 70, 70, 90, 100]
                    top_rho = 2400
                    bottom_rho = 1200
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových tvárnic s hutným kamenivem skupiny 1'
                    token_tab = '6.3.1, číslo řádku 1.3 respektive 1.4'
                    token_type = 'EI'

                if Specif2[i] == 'Beton nenosné skupina 2 pórovité kamenivo 240<p<1200':
                    interpolate = []
                    top = [50, 70, 100, 100, 140, 200]
                    bottom = [50, 50, 70, 90, 140, 140]
                    top_rho = 1200
                    bottom_rho = 240
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových tvárnic s pórovitým kamenivem skupiny 2'
                    token_tab = '6.3.1, číslo řádku 2.1 respektive 2.2'
                    token_type = 'EI'

                if Specif2[i] == 'Beton nenosné skupina 2 hutné kamenivo 720<p<1650':
                    interpolate = []
                    top = [50, 70, 100, 100, 200, 200]
                    bottom = [50, 50, 70, 70, 140, 140]
                    top_rho = 1650
                    bottom_rho = 720
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových tvárnic s hutným kamenivem skupiny 2'
                    token_tab = '6.3.1, číslo řádku 2.3 respektive 2.4'
                    token_type = 'EI'

                if Specif2[i] == 'Beton nenosné skupina 3 hutné kamenivo 480<p<1000':
                    interpolate = [100, 150, 150, 200, 1000, 1000]
                    top_rho = 1000
                    bottom_rho = 480
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových tvárnic s hutným kamenivem skupiny 3'
                    token_tab = '6.3.1, číslo řádku 3.3'
                    token_type = 'EI'

                if Specif2[i] == 'Beton nosné skupina 1 pórovité kamenivo 400<p<1600':
                    interpolate = []
                    top = [170, 170, 170, 170, 190, 240]
                    bottom = [140, 140, 140, 140, 170, 190]
                    top_rho = 1600
                    bottom_rho = 400
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových nosných tvárnic s pórovitým kamenivem skupiny 1'
                    token_tab = '6.3.2, číslo řádku 1.1 respektive 1.2'
                    token_type = 'REI'

                if Specif2[i] == 'Beton nosné skupina 1 hutné kamenivo 1400<p<2000':
                    interpolate = []
                    top = [170, 170, 170, 170, 190, 240]
                    bottom = [140, 140, 140, 140, 170, 190]
                    top_rho = 2000
                    bottom_rho = 1400
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových nosných tvárnic s hutným kamenivem skupiny 1'
                    token_tab = '6.3.2, číslo řádku 1.3 respektive 1.4'
                    token_type = 'REI'

                if Specif2[i] == 'Beton nosné skupina 2 pórovité kamenivo 240<p<1200':
                    interpolate = []
                    top = [170, 170, 170, 170, 190, 240]
                    bottom = [140, 140, 140, 140, 170, 190]
                    top_rho = 1200
                    bottom_rho = 240
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových nosných tvárnic s pórovitým kamenivem skupiny 2'
                    token_tab = '6.3.2, číslo řádku 2.1 respektive 2.2'
                    token_type = 'REI'

                if Specif2[i] == 'Beton nosné skupina 2 hutné kamenivo 720<p<1650':
                    interpolate = []
                    top = [170, 170, 170, 170, 190, 240]
                    bottom = [140, 140, 140, 140, 170, 190]
                    top_rho = 1650
                    bottom_rho = 720
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových nosných tvárnic s hutným kamenivem skupiny 2'
                    token_tab = '6.3.2, číslo řádku 2.3 respektive 2.4'
                    token_type = 'REI'

                if Specif2[i] == 'Beton nosné skupina 3 hutné kamenivo 480<p<1000':
                    interpolate = [140, 140, 140, 140, 200, 200]
                    top_rho = 1000
                    bottom_rho = 480
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z betonových nosných tvárnic s hutným kamenivem skupiny 3'
                    token_tab = '6.3.2, číslo řádku 3.3'
                    token_type = 'REI'

                if Specif2[i] == 'Pórobeton nenosné skupina 1S a 1 350<p<500':
                    interpolate = []
                    top = [65, 70, 75, 100, 100, 150]
                    bottom = [50, 60, 60, 60, 90, 100]
                    top_rho = 500
                    bottom_rho = 350
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pórobetonových tvárnic'
                    token_tab = '6.4.1, číslo řádku 1.1 respektive 1.2'
                    token_type = 'EI'

                if Specif2[i] == 'Pórobeton nenosné skupina 1S a 1 500<p<1000':
                    interpolate = []
                    top = [60, 60, 70, 100, 100, 150]
                    bottom = [50, 60, 60, 60, 90, 100]
                    top_rho = 1000
                    bottom_rho = 500
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z pórobetonových tvárnic'
                    token_tab = '6.4.1, číslo řádku 1.3 respektive 1.4'
                    token_type = 'EI'

                if Specif2[i] == 'Pórobeton nosné skupina 1S a 1 350<p<500':
                    interpolate = []
                    top = [115, 115, 140, 200, 225, 300]
                    bottom = [115, 115, 115, 200, 125, 240]
                    top_rho = 500
                    bottom_rho = 350
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z nosných pórobetonových tvárnic'
                    token_tab = '6.4.2, číslo řádku 1.1 respektive 1.2'
                    token_type = 'REI'

                if Specif2[i] == 'Pórobeton nosné skupina 1S a 1 500<p<1000':
                    interpolate = []
                    top = [100, 100, 150, 170, 200, 240]
                    bottom = [100, 100, 100, 150, 170, 200]
                    top_rho = 1000
                    bottom_rho = 500
                    ' Vyzkoušení zda je rho v intervalu'
                    if a[i] > top_rho or a[i] < bottom_rho:
                        raise NameError('!!! Mimo rozsah tabulkek !!!')
                    for n in range(0, len(top)):
                        intt = top[n] + (a[i]-bottom_rho) *\
                               ((bottom[n]-top[n])/(top_rho-bottom_rho))
                        interpolate.append(intt)
                        myArr = np.asarray(interpolate)
                    check = cihlyPO(b, interpolate, i, myArr)
                    token = 'z nosných pórobetonových tvárnic'
                    token_tab = '6.4.2, číslo řádku 1.3 respektive 1.4'
                    token_type = 'REI'

            doc.append(NoEscape(r'\textbf{%s}: ' % Nazev_kce[i]))
            if Specif2[i] in dictionary_text:
                doc.append(dictionary_text[Specif2[i]])
            if 'PP' in Nazev_kce[i] or 'podzemní' in Nazev_kce[i]\
               or 'Podzemní' in Nazev_kce[i]:
                if check_PP[spb_max] > check:
                    raise NameError('!!! Konstrukce nevyhoví na požární odolnost !!!')
                else:
                    add_str = 'v podzemním podlaží'
            if 'NP' in Nazev_kce[i] or 'nadzemní' in Nazev_kce[i]\
               or 'Nadzemní' in Nazev_kce[i]:
                add_str = 'v nadzemním podlaží'
                if check_NP[spb_max] > check:
                    raise NameError('!!! Konstrukce nevyhoví na požární odolnost !!!')
                else:
                    if i != 0:
                        if b[i] in save_b and Specif2[i] in save_var:
                            doc.append('Požární stěny {} jsou stejně jako v podzemním podlaží tvořeny {} o tloušťce {} mm s vyhovující požární odolností {} {} DP1. '.format(add_str, token, b[i], token_type, check))
                            doc.append(NoEscape(r'\newline \newline'))
            if i == 0 or (b[i] not in save_b and Specif2[i] not in save_var):
                if Specif2[i] == 'Neznámé cihly':
                    doc.append('Požární stěna je tvořena neznámým druhem pálených cihel. Stěna o tloušťce {} mm vykazuje dle tabulkového hodnocení tab: {} (Zoufal a spol.) i při uvažování cihel skupiny 4, tedyskupiny, která z požárního hlediska vykazuje nejhorší požární odolnost, vyhovující odolnost {} {} DP1. '.format(b[i], token_tab, token_type, check))
                elif Specif2[i] in dictionary and Specif2[i] != 'Neznámé cihly':
                    doc.append('Požární stěny {} jsou tvořeny {} o tloušťce {} mm. Dle tabulkového hodnocení tab: {} (viz Zoufal a spol.) má konstrukce požární odolnost {} {} DP1, což je vyhovující. '.format(add_str, token, b[i], token_tab, token_type, check))
            doc.append(NoEscape(r'\newline \newline'))
            save_var.append(Specif2[i])
            save_b.append(b[i])
    os.chdir(vystup_dir)
    doc.generate_pdf("E.1", clean_tex=False)
