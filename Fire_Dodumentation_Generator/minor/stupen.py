import openpyxl
import numpy as np
file = 'SPB.xlsx'
wb = openpyxl.load_workbook(file, data_only=True)
SPB = wb["SPB"]


def chunks(l, n):  # This definition splits the list into desired chunks
    # For item i in a range that is a length of l,
    for i in range(0, len(l), n):
        # Create an index range for l of n items:
        yield l[i:i+n]


def spb_def(h_p, type_sys, p, a, podlazi, n_uniques):
    ###########################################################
    ''' Zatřídění do stupně požární bezpenčosti'''
    # Volání proměnných
    # Info = wb.get_sheet_by_name('Info')  # Opens the sheet called Info
    # Opens the sheet called Info
    SPB_res = ['I.', 'II.', 'III.', 'IV.', 'V.', 'VI.', 'VII.']
    SPB_result = []
    if type_sys == 'nehořlavý':
        data_SPB = []  # empty list for storing data
        data_p = [15, 30, 45, 60, 90, 120, 1000]
        for row in SPB.iter_rows(min_row=2, max_row=8,
                                 min_col=2, max_col=8):
            for cell in row:
                data_SPB.append(cell.value)
        data_SPB = list(chunks(data_SPB, 7))

    if type_sys == 'smíšený':
        data_SPB = []  # empty list for storing data
        data_p = [10, 25, 35, 50, 75, 100, 1000]
        for row in SPB.iter_rows(min_row=9, max_row=15,
                                 min_col=2, max_col=6):
            for cell in row:
                data_SPB.append(cell.value)
        data_SPB = list(chunks(data_SPB, 5))

    if type_sys == 'hořlavý':
        data_SPB = []  # empty list for storing data
        data_p = [10, 20, 30, 40, 60, 80, 1000]
        for row in SPB.iter_rows(min_row=16, max_row=22,
                                 min_col=2, max_col=6):
            for cell in row:
                data_SPB.append(cell.value)
        data_SPB = list(chunks(data_SPB, 5))

    h_p_init = h_p
    for i in range(0, len(p)):
        if isinstance(n_uniques, list):
            h_p = h_p_init
            if n_uniques[i][0] == 'P' and h_p_init > 6.0:
                if n_uniques[i][1] == '1':
                        h_p = 22.5
                if n_uniques[i][1] == '2' or n_uniques[i][1] == '3' or\
                   n_uniques[i][1] == '4' or n_uniques[i][1] == '5' or\
                   n_uniques[i][1] == '6' or n_uniques[i][1] == '7':
                        h_p = 30.0

            if n_uniques[i][0] == 'P' and h_p_init <= 6.0:
                if n_uniques[i][1] == '1':
                        h_p = h_p
                if n_uniques[i][1] == '2' or n_uniques[i][1] == '3' or\
                   n_uniques[i][1] == '4' or n_uniques[i][1] == '5' or\
                   n_uniques[i][1] == '6' or n_uniques[i][1] == '7':
                    if h_p <= 6.0:
                        h_p = 12

        if isinstance(n_uniques, str):
            h_p = h_p_init
            if n_uniques[0] == 'P' and h_p_init > 6.0:
                if n_uniques[1] == '1':
                        h_p = 22.5
                if n_uniques[1] == '2' or n_uniques[1] == '3' or\
                   n_uniques[1] == '4' or n_uniques[1] == '5' or\
                   n_uniques[1] == '6' or n_uniques[1] == '7':
                        h_p = 30.0

            if n_uniques[0] == 'P' and h_p_init <= 6.0:
                if n_uniques[1] == '1':
                        h_p = h_p
                if n_uniques[1] == '2' or n_uniques[1] == '3' or\
                   n_uniques[1] == '4' or n_uniques[1] == '5' or\
                   n_uniques[1] == '6' or n_uniques[1] == '7':
                    if h_p <= 6.0:
                        h_p = 12

        if type_sys == 'nehořlavý':
            if p[i] in data_p:
                val_data_p = p[i]
            if p[i] not in data_p:
                val_data_p = min(filter(lambda x: x > p[i], data_p))
            # print(val_data_p)
            idx_SPB = data_p.index(val_data_p)
            temp_SPB = data_SPB[idx_SPB]
            if a[i] <= 1.1 and h_p == 0 and idx_SPB == 6:
                idx_SPB_h = 2
            if a[i] > 1.1 and h_p == 0 and idx_SPB == 6:
                idx_SPB_h = 3
            if a[i] <= 1.1 and h_p == 0 and idx_SPB == 5:
                idx_SPB_h = 1
            if a[i] > 1.1 and h_p == 0 and idx_SPB == 5:
                idx_SPB_h = 2
            if a[i] <= 1.1 and h_p == 0 and idx_SPB == 4:
                idx_SPB_h = 0
            if a[i] > 1.1 and h_p == 0 and idx_SPB == 4:
                idx_SPB_h = 1
            if h_p == 0 and idx_SPB == 3:
                idx_SPB_h = 0
            if h_p == 0 and idx_SPB == 2:
                idx_SPB_h = 0
            if h_p == 0 and idx_SPB == 1:
                idx_SPB_h = 0
            if h_p > 0:
                if h_p in temp_SPB:
                    val_data_h = h_p
                    idx_SPB_h = temp_SPB.index(val_data_h)
                if h_p not in temp_SPB:
                    if h_p > max(temp_SPB) and idx_SPB < 5:
                        idx_SPB_h = (temp_SPB.index(max(temp_SPB))+1)
                    if h_p <= max(temp_SPB) and idx_SPB < 5:
                        val_data_h = min(filter(lambda x: x > h_p, temp_SPB))
                        idx_SPB_h = temp_SPB.index(val_data_h)
                    if h_p > max(temp_SPB) and idx_SPB >= 5:
                        raise NameError('!!! Mimo rozsah: sniz h_p nabo p_v !!!')
                    if h_p <= max(temp_SPB) and idx_SPB >= 5:
                        val_data_h = min(filter(lambda x: x > h_p, temp_SPB))
                        idx_SPB_h = temp_SPB.index(val_data_h)
            SPB_result.append(SPB_res[idx_SPB_h])

        if type_sys == 'smíšený':
            if p[i] in data_p:
                val_data_p = p[i]
            if p[i] not in data_p:
                val_data_p = min(filter(lambda x: x > p[i], data_p))
            idx_SPB = data_p.index(val_data_p)
            temp_SPB = data_SPB[idx_SPB]
            if h_p == 0 and idx_SPB == 6:
                idx_SPB_h = 2
            if h_p == 0 and idx_SPB == 5:
                idx_SPB_h = 1
            if h_p == 0 and idx_SPB == 4:
                idx_SPB_h = 1
            if a[i] <= 1.1 and h_p == 0 and idx_SPB == 3:
                idx_SPB_h = 0
            if a[i] > 1.1 and h_p == 0 and idx_SPB == 3:
                idx_SPB_h = 1
            if h_p == 0 and idx_SPB == 2:
                idx_SPB_h = 0
            if h_p == 0 and idx_SPB == 1:
                idx_SPB_h = 0
            if h_p > 0:
                if h_p in temp_SPB:
                    val_data_h = h_p
                    idx_SPB_h = temp_SPB.index(val_data_h)
                if h_p not in temp_SPB:
                    if h_p <= max(temp_SPB):
                        val_data_h = min(filter(lambda x: x > h_p, temp_SPB))
                        idx_SPB_h = temp_SPB.index(val_data_h)
                    if h_p > max(temp_SPB):
                        raise NameError('!!! Mimo rozsah: sniz h_p nabo p_v !!!')
            SPB_result.append(SPB_res[idx_SPB_h])

        if type_sys == 'hořlavý':
            if p[i] in data_p:
                val_data_p = p[i]
            if p[i] not in data_p:
                val_data_p = min(filter(lambda x: x > p[i], data_p))
            idx_SPB = data_p.index(val_data_p)
            temp_SPB = data_SPB[idx_SPB]
            if a[i] <= 1.1 and h_p == 0 and idx_SPB == 6:
                idx_SPB_h = 2
            if a[i] > 1.1 and h_p == 0 and idx_SPB == 6:
                idx_SPB_h = 3
            if a[i] <= 1.1 and h_p == 0 and idx_SPB == 5:
                idx_SPB_h = 1
            if a[i] > 1.1 and h_p == 0 and idx_SPB == 5:
                idx_SPB_h = 2
            if h_p == 0 and idx_SPB == 4:
                idx_SPB_h = 1
            if a[i] <= 1.1 and h_p == 0 and idx_SPB == 3:
                idx_SPB_h = 0
            if a[i] > 1.1 and h_p == 0 and idx_SPB == 3:
                idx_SPB_h = 1
            if h_p == 0 and idx_SPB == 2:
                idx_SPB_h = 0
            if h_p == 0 and idx_SPB == 1:
                idx_SPB_h = 0
            if h_p > 0:
                if h_p in temp_SPB:
                    val_data_h = h_p
                    idx_SPB_h = temp_SPB.index(val_data_h)
                if h_p not in temp_SPB:
                    if h_p <= max(temp_SPB):
                        val_data_h = min(filter(lambda x: x > h_p, temp_SPB))
                        idx_SPB_h = temp_SPB.index(val_data_h)
                        print(idx_SPB_h)
                    if h_p > max(temp_SPB):
                        raise NameError('!!! Mimo rozsah: sniz h_p nabo p_v !!!')
            SPB_result.append(SPB_res[idx_SPB_h])
    return SPB_result


def rozmer(a, type_sys, podlazi, h_p):
    data_a = [0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0, 1.1, 1.2, 1.3]
    if type_sys == 'nehořlavý':
        data_maxPU = []  # empty list for storing data
        for row in SPB.iter_rows(min_row=23, max_row=33,
                                 min_col=1, max_col=9):
            for cell in row:
                data_maxPU.append(cell.value)
        data_maxPU = list(chunks(data_maxPU, 9))

    delka_PU = np.zeros(len(a))
    sirka_PU = np.zeros(len(a))
    for i in range(0, len(a)):
        if a[i] in data_a:
            idx = data_a.index(a[i])
            for k in range(0, len(data_maxPU)):
                if data_maxPU[k][0] == a[i]:
                    if podlazi == 1:
                        delka_PU[i] = data_maxPU[idx][1]
                        sirka_PU[i] = data_maxPU[idx][2]
                    if podlazi > 1 and h_p <= 22.5:
                        delka_PU[i] = data_maxPU[idx][3]
                        sirka_PU[i] = data_maxPU[idx][4]
                    if podlazi > 1 and h_p > 22.5 and h_p <= 45:
                        delka_PU[i] = data_maxPU[idx][5]
                        sirka_PU[i] = data_maxPU[idx][6]
                    if podlazi > 1 and h_p > 45:
                        delka_PU[i] = data_maxPU[idx][7]
                        sirka_PU[i] = data_maxPU[idx][8]
        if a[i] not in data_a:
            if a[i] <= 0.3:
                val_data_min = min(filter(lambda x: x > a[i], data_a))
                idx_PU = data_a.index(val_data_min)
                if podlazi == 1:
                    delka_PU[i] = data_maxPU[idx_PU][1]
                    sirka_PU[i] = data_maxPU[idx_PU][2]
                if podlazi > 1 and h_p <= 22.5:
                    delka_PU[i] = data_maxPU[idx_PU][3]
                    sirka_PU[i] = data_maxPU[idx_PU][4]
                if podlazi > 1 and h_p > 22.5 and h_p <= 45:
                    delka_PU[i] = data_maxPU[idx_PU][5]
                    sirka_PU[i] = data_maxPU[idx_PU][6]
                if podlazi > 1 and h_p > 45:
                    delka_PU[i] = data_maxPU[idx_PU][7]
                    sirka_PU[i] = data_maxPU[idx_PU][8]
            if 0.3 < a[i] < 1.3:
                val_data_max = min(filter(lambda x: x > a[i], data_a))
                val_data_min = max(filter(lambda x: x < a[i], data_a))
                idx_minPU = data_a.index(val_data_min)
                idx_maxPU = data_a.index(val_data_max)
                if podlazi == 1:
                    delka_min = data_maxPU[idx_minPU][1]
                    sirka_min = data_maxPU[idx_minPU][2]
                    delka_max = data_maxPU[idx_maxPU][1]
                    sirka_max = data_maxPU[idx_maxPU][2]
                if podlazi > 1 and h_p <= 22.5:
                    delka_min = data_maxPU[idx_minPU][3]
                    sirka_min = data_maxPU[idx_minPU][4]
                    delka_max = data_maxPU[idx_maxPU][3]
                    sirka_max = data_maxPU[idx_maxPU][4]
                if podlazi > 1 and h_p > 22.5 and h_p <= 45:
                    delka_min = data_maxPU[idx_minPU][5]
                    sirka_min = data_maxPU[idx_minPU][6]
                    delka_max = data_maxPU[idx_maxPU][5]
                    sirka_max = data_maxPU[idx_maxPU][6]
                if podlazi > 1 and h_p > 45:
                    delka_min = data_maxPU[idx_minPU][7]
                    sirka_min = data_maxPU[idx_minPU][8]
                    delka_max = data_maxPU[idx_maxPU][7]
                    sirka_max = data_maxPU[idx_maxPU][8]
                delka_PU[i] = delka_min + (a[i] - val_data_min) * \
                                          ((delka_max - delka_min) /
                                           (val_data_max - val_data_min))
                sirka_PU[i] = sirka_min + (a[i] - val_data_min) * \
                                          ((sirka_max - sirka_min) /
                                           (val_data_max - val_data_min))
            if a[i] >= 1.3:
                val_data_max = max(filter(lambda x: x < a[i], data_a))
                idx_PU = data_a.index(val_data_max)
                if podlazi == 1:
                    delka_PU[i] = data_maxPU[idx_PU][1]
                    sirka_PU[i] = data_maxPU[idx_PU][2]
                if podlazi > 1 and h_p <= 22.5:
                    delka_PU[i] = data_maxPU[idx_PU][3]
                    sirka_PU[i] = data_maxPU[idx_PU][4]
                if podlazi > 1 and h_p > 22.5 and h_p <= 45:
                    delka_PU[i] = data_maxPU[idx_PU][5]
                    sirka_PU[i] = data_maxPU[idx_PU][6]
                if podlazi > 1 and h_p > 45:
                    delka_PU[i] = data_maxPU[idx_PU][7]
                    sirka_PU[i] = data_maxPU[idx_PU][8]
    return delka_PU, sirka_PU
